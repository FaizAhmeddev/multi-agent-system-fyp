"""
UTILS/FILE_PARSER.PY
=====================
Parse PDF, DOCX, CSV, JSON, XLSX files into plain text.
Used by finance_drive_reader and documents_agent.
"""

import os


def parse_file(path: str) -> list[str]:
    """
    Parse a file at `path` and return a list of text chunks.
    Returns empty list on any error (never raises).
    """
    if not path or not os.path.exists(path):
        return []

    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".pdf":
            return _parse_pdf(path)
        elif ext == ".docx":
            return _parse_docx(path)
        elif ext == ".csv":
            return _parse_csv(path)
        elif ext == ".json":
            return _parse_json(path)
        elif ext in (".xlsx", ".xls"):
            return _parse_excel(path)
        elif ext == ".txt":
            return _parse_txt(path)
        else:
            return []
    except Exception as e:
        return [f"[Parse error for {os.path.basename(path)}: {e}]"]


def parse_file_to_text(path: str) -> str:
    """Convenience: return all chunks joined as a single string."""
    chunks = parse_file(path)
    return "\n".join(chunks)


# ── private helpers ────────────────────────────────────────────────────────────

def _parse_pdf(path: str) -> list[str]:
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return [text] if text.strip() else ["[PDF has no extractable text]"]
    except ImportError:
        # fallback to pypdf
        from pypdf import PdfReader
        reader = PdfReader(path)
        text = ""
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
        return [text] if text.strip() else ["[PDF has no extractable text]"]


def _parse_docx(path: str) -> list[str]:
    from docx import Document
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [text] if text.strip() else ["[DOCX is empty]"]


def _parse_csv(path: str) -> list[str]:
    import pandas as pd
    df = pd.read_csv(path)
    # Return header + all rows as individual strings for large files
    rows = [",".join(str(v) for v in row) for row in df.values]
    header = ",".join(df.columns.tolist())
    return [header] + rows


def _parse_json(path: str) -> list[str]:
    import json
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return [json.dumps(data, indent=2)]


def _parse_excel(path: str) -> list[str]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        results = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
            if rows:
                results.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        return results if results else ["[Excel file is empty]"]
    except Exception:
        import pandas as pd
        df = pd.read_excel(path)
        return [df.to_string()]


def _parse_txt(path: str) -> list[str]:
    with open(path, encoding="utf-8", errors="replace") as f:
        return [f.read()]
